"""
Run 006 - Generate Trajectory Comparison Tables (Excel Format)
===============================================================

Generate Excel file with trajectory comparison tables for all seeds.
Each seed is in a separate sheet.

Author: zhanghl
Date: 2026-04-09
"""

import os
import sys
import pandas as pd
import numpy as np
from typing import List, Tuple
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# Import gene categories from ddc
sys.path.insert(0, '/home/zhanghl/projects/ddc_github/src')
import ddc
from ddc import GENE_CATEGORIES

# ==========================================
# Configuration
# ==========================================
BASE_DIR = '/home/zhanghl/projects/ddc_github/runs/run_006'
DATA_DIR = os.path.join(BASE_DIR, 'data')
RESULTS_DIR = os.path.join(BASE_DIR, 'results')

TRAJ_DIR = os.path.join(DATA_DIR, 'trajectories')
METRICS_DIR = os.path.join(DATA_DIR, 'response_metrics')
OUTPUT_DIR = os.path.join(DATA_DIR, 'trajectory_tables')

BASELINE_TSV = os.path.join(TRAJ_DIR, 'baseline.tsv')
PERTURBED_TSV = os.path.join(TRAJ_DIR, 'perturbed.tsv')
RESPONSE_TSV = os.path.join(METRICS_DIR, 'response_summary.tsv')

# Time window after perturbation
WINDOW_AFTER = 10   # Show perturbation time and 10 steps after (total 11 time points)


def setup_output_dir():
    """Create output directory if not exists."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)


def load_response_summary() -> pd.DataFrame:
    """Load response summary to get seed info and T_perturb."""
    return pd.read_csv(RESPONSE_TSV, sep='\t')


def load_trajectory_data(tsv_path: str) -> pd.DataFrame:
    """Load trajectory TSV file."""
    return pd.read_csv(tsv_path, sep='\t')


def extract_seed_trajectory(df: pd.DataFrame, seed: int) -> pd.DataFrame:
    """Extract trajectory data for a specific seed."""
    return df[df['seed'] == seed].copy()


def get_time_window(T_perturb: int, after: int) -> List[int]:
    """Get time points from perturbation time onwards."""
    t_start = T_perturb
    t_end = T_perturb + after
    return list(range(t_start, t_end + 1))


def get_gene_type(gene_id: int) -> str:
    """Get gene type from GENE_CATEGORIES."""
    for category, subcategories in GENE_CATEGORIES.items():
        for subcat_name, gene_list in subcategories.items():
            if gene_id in gene_list:
                return subcat_name
    return "Unknown"


def create_comparison_table(
    baseline_df: pd.DataFrame,
    perturbed_df: pd.DataFrame,
    T_perturb: int,
    target_gene: int,
    perturbation_type: str,
    window_after: int = 9
) -> pd.DataFrame:
    """
    Create a comparison table with baseline and perturbed as separate rows.
    
    Format:
        Gene | Type      | t=T | t=T+1 | t=T+2 | ... | t=T+9
        0    | baseline  | ... | ...   | ...   | ... | ...
        0    | perturbed | ... | ...   | ...   | ... | ...
        1    | baseline  | ... | ...   | ...   | ... | ...
        1    | perturbed | ... | ...   | ...   | ... | ...
    """
    time_points = get_time_window(T_perturb, window_after)
    
    # Get unique gene IDs
    genes = sorted(baseline_df['gene_id'].unique())
    
    # Build table data - two rows per gene
    data = []
    for gene in genes:
        # Baseline row
        row_baseline = {'gene_id': gene, 'type': 'baseline'}
        # Perturbed row
        row_perturbed = {'gene_id': gene, 'type': 'perturbed'}
        
        for t in time_points:
            # Get baseline P at time t
            baseline_p = baseline_df[(baseline_df['time'] == t) & 
                                      (baseline_df['gene_id'] == gene)]['P'].values
            baseline_p_val = baseline_p[0] if len(baseline_p) > 0 else np.nan
            
            # Get perturbed P at time t
            perturbed_p = perturbed_df[(perturbed_df['time'] == t) & 
                                        (perturbed_df['gene_id'] == gene)]['P'].values
            perturbed_p_val = perturbed_p[0] if len(perturbed_p) > 0 else np.nan
            
            # Add columns with time as header
            row_baseline[f't={t}'] = baseline_p_val
            row_perturbed[f't={t}'] = perturbed_p_val
        
        data.append(row_baseline)
        data.append(row_perturbed)
    
    df_table = pd.DataFrame(data)
    
    # Add gene_type column
    df_table.insert(0, 'gene_type', df_table['gene_id'].apply(get_gene_type))
    
    return df_table


def write_sheet_with_merged_cells(ws, df: pd.DataFrame, sheet_name: str, seed: int, regime: str, T_perturb: int, target_gene: int, pert_type: str):
    """Write dataframe to worksheet with merged cells for gene_type and gene_id."""
    # Write compact header (2 lines)
    ws['A1'] = f"Trajectory Comparison Table - Seed {seed} ({regime})"
    ws['A1'].font = Font(bold=True, size=14, name='Calibri')
    
    time_points = [col for col in df.columns if col.startswith('t=')]
    t_start = time_points[0].replace('t=', '')
    t_end = time_points[-1].replace('t=', '')
    ws['A2'] = f"Perturbation Time: t={T_perturb} | Target Gene: {target_gene} | Perturbation Type: {pert_type} | Time Points: {t_start} to {t_end}"
    ws['A2'].font = Font(size=12, name='Calibri')
    
    # Write dataframe starting from row 3 (no empty line after header)
    thin_border = Border(bottom=Side(style='thin'))
    
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=3):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            
            # Get gene_id for this row (column B, index 2)
            gene_id = ws.cell(row=r_idx, column=2).value
            # Get type for this row (column C, index 3)
            row_type = ws.cell(row=r_idx, column=3).value
            
            # Format header row
            if r_idx == 3:
                cell.font = Font(bold=True, size=12, name='Calibri')
                cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            else:
                # Default font
                cell.font = Font(name='Calibri')
                
                # Bold the target gene rows (all columns from gene_id onwards)
                if gene_id is not None and gene_id == target_gene:
                    if c_idx >= 2:  # gene_id and all subsequent columns
                        cell.font = Font(bold=True, name='Calibri')
                
                # Color the perturbed value at T_perturb time (lighter deep red)
                perturbed_col = f't={T_perturb}'
                if c_idx > 3 and gene_id == target_gene:  # Time columns start at column 4
                    col_name = df.columns[c_idx - 1]  # Column name in df
                    if col_name == perturbed_col:
                        cell.font = Font(bold=True, color="C41E3A", name='Calibri')  # Lighter deep red (Cardinal)
                
                # Format numbers
                if isinstance(value, float):
                    cell.number_format = '0.0000'
                
                # Add bottom border between different genes (only on perturbed rows)
                if row_type == 'perturbed':
                    cell.border = thin_border
    
    # Merge cells for gene_type column
    merge_cells_in_column(ws, 1, start_row=4)  # gene_type is column 1
    
    # Merge cells for gene_id column
    merge_cells_in_column(ws, 2, start_row=4)  # gene_id is column 2
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 15  # gene_type
    ws.column_dimensions['B'].width = 10  # gene_id
    ws.column_dimensions['C'].width = 12  # type
    for i, col in enumerate(df.columns[3:], start=4):  # time columns
        ws.column_dimensions[ws.cell(row=3, column=i).column_letter].width = 10
    
    # Align cells
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')


def merge_cells_in_column(ws, col_idx: int, start_row: int):
    """Merge consecutive cells with same value in a column."""
    row = start_row
    while row <= ws.max_row:
        # Find range of rows with same value
        current_value = ws.cell(row=row, column=col_idx).value
        if current_value is None:
            break
        
        end_row = row
        while end_row + 1 <= ws.max_row and ws.cell(row=end_row + 1, column=col_idx).value == current_value:
            end_row += 1
        
        # Merge if more than one row
        if end_row > row:
            ws.merge_cells(start_row=row, start_column=col_idx, end_row=end_row, end_column=col_idx)
        
        row = end_row + 1


def generate_excel_file():
    """Generate Excel file with all seeds as separate sheets."""
    print("=" * 60)
    print("Generating Trajectory Comparison Tables (Excel)")
    print("=" * 60)
    
    setup_output_dir()
    
    # Load data
    print("\nLoading data...")
    response_df = load_response_summary()
    baseline_df = load_trajectory_data(BASELINE_TSV)
    perturbed_df = load_trajectory_data(PERTURBED_TSV)
    
    print(f"  - Response summary: {len(response_df)} seeds")
    print(f"  - Baseline trajectories: {len(baseline_df)} records")
    print(f"  - Perturbed trajectories: {len(perturbed_df)} records")
    
    # Create Excel workbook
    wb = Workbook()
    # Remove default sheet
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Process each seed
    print("\nGenerating sheets...")
    for _, row in response_df.iterrows():
        seed = int(row['seed'])
        regime = row['regime']
        T_perturb = int(row['T_perturb'])
        target_gene = int(row['target_gene'])
        pert_type = row['perturbation_type']
        
        print(f"\n  Seed {seed} ({regime}, {pert_type} at t={T_perturb}, gene={target_gene}):")
        
        # Extract seed-specific data
        seed_baseline = extract_seed_trajectory(baseline_df, seed)
        seed_perturbed = extract_seed_trajectory(perturbed_df, seed)
        
        print(f"    - Baseline records: {len(seed_baseline)}")
        print(f"    - Perturbed records: {len(seed_perturbed)}")
        
        # Create comparison table
        table = create_comparison_table(
            seed_baseline,
            seed_perturbed,
            T_perturb,
            target_gene,
            pert_type,
            WINDOW_AFTER
        )
        
        print(f"    - Table shape: {table.shape}")
        
        # Create sheet
        sheet_name = f"seed_{seed}"
        ws = wb.create_sheet(title=sheet_name)
        
        # Write data with merged cells
        write_sheet_with_merged_cells(ws, table, sheet_name, seed, regime, T_perturb, target_gene, pert_type)
        
        print(f"    - Sheet created: {sheet_name}")
    
    # Save Excel file
    output_path = os.path.join(OUTPUT_DIR, 'run_006_trajectory_tables.xlsx')
    wb.save(output_path)
    
    print("\n" + "=" * 60)
    print("Done! Excel file saved to:")
    print(f"  {output_path}")
    print("=" * 60)


if __name__ == '__main__':
    generate_excel_file()
